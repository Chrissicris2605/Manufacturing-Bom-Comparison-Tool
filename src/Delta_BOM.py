#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from pathlib import Path
import sys
import re
from collections import defaultdict, deque
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# -----------------------
# PySide6 (GUI)
# -----------------------
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QIcon
from PySide6.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QTextEdit, QPushButton, QCheckBox,
    QFileDialog, QHBoxLayout, QVBoxLayout
)
from PySide6.QtGui import QPixmap
from pathlib import Path
from tempfile import NamedTemporaryFile
import pandas as pd

# ===========================
#  Temas (QSS)
# ===========================
DARK_QSS = """
QWidget { background-color: #1e1e2f; color: #f0f0f0; font-size: 12pt; }
QLabel  { color: #f0f0f0; }
QLineEdit, QTextEdit { background: #1c1c2c; border: 1px solid #555; padding: 6px; }
QPushButton { background-color: #0a84ff; color: white; padding: 8px 12px; border-radius: 6px; }
QPushButton:hover { background-color: #006ddc; }
QCheckBox { spacing: 8px; color: #f0f0f0; }
"""

LIGHT_QSS = """
QWidget { background-color: #ffffff; color: #002244; font-size: 12pt; }
QLabel  { color: #66a6ff; }             /* labels em azul claro */
QLineEdit, QTextEdit { background: #ffffff; color: #002244; border: 1px solid #aab; padding: 6px; }
QPushButton { background-color: #003e92; color: white; padding: 8px 12px; border-radius: 6px; }
QPushButton:hover { background-color: #002f70; }
QCheckBox { spacing: 8px; color: #002244; }
/* Checkbox visível no modo claro */
QCheckBox::indicator {
    width: 16px;
    height: 16px;
    border: 1px solid #003e92;
    border-radius: 3px;
    background: white;
}
QCheckBox::indicator:checked {
    background-color: #003e92;
    image: url("");
}
"""

# ===========================
#  Cores/estilos planilha
# ===========================
PINK  = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
BLUE  = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
YELL  = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
RED   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
GREEN = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
DARK_BLUE = PatternFill(start_color="000080", end_color="000080", fill_type="solid")

WHITE_BOLD = Font(color="FFFFFF", bold=True)
BLACK_BOLD = Font(color="000000", bold=True)
BOLD = Font(bold=True)

# Índices 1-based (antes de inserções)
COL_A = 1
COL_C = 3
COL_D = 4
COL_E = 5
COL_F = 6
COL_J = 10
COL_K = 11
COL_L = 12
COL_M = 13
COL_N = 14
COL_O = 15
COL_Q = 16  # status

NBSP = "\u00A0"

def clean_str(s: str) -> str:
    if s is None:
        return ""
    s = str(s).replace(NBSP, " ").strip()
    return re.sub(r"\s+", " ", s)

def parse_number_relaxed(s: str):
    if s is None:
        return None
    s = str(s).strip().replace(NBSP, " ")
    if s == "":
        return None
    if s.count(",") == 1 and s.count(".") >= 1:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except:
        return None

def cellv(ws, r, c):
    v = ws.cell(row=r, column=c).value
    return "" if v is None else str(v)

def celln(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return parse_number_relaxed(v)

def equal_L(ws, r1, r2, COL_L, tol=1e-6):
    v1 = celln(ws, r1, COL_L)
    v2 = celln(ws, r2, COL_L)
    if v1 is not None and v2 is not None:
        return abs(v1 - v2) <= tol
    return clean_str(cellv(ws, r1, COL_L)) == clean_str(cellv(ws, r2, COL_L))

def auto_width(ws, max_width=80):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        m = 0
        for cell in col:
            s = "" if cell.value is None else str(cell.value)
            m = max(m, len(s))
        ws.column_dimensions[letter].width = min(m+2, max_width)
        
def xlsb_to_xlsx(xlsb_path: Path, sheet_name: str) -> Path:
    """
    Lê um .xlsb (somente valores) e grava um .xlsx temporário sem formatos/mesclas,
    preservando o conteúdo das células na mesma posição.
    """
    df = pd.read_excel(xlsb_path, sheet_name=sheet_name, engine="pyxlsb", header=None)
    tmp = NamedTemporaryFile(suffix=".xlsx", delete=False)
    # header=None para não perder as 7 primeiras linhas do seu layout (títulos na linha 7, etc.)
    df.to_excel(tmp.name, index=False, header=False, engine="openpyxl")
    return Path(tmp.name)


def ensure_xlsx(path: Path, sheet_name: str) -> Path:
    """
    Se o arquivo for .xlsb, converte para .xlsx temporário.
    Caso contrário, retorna o próprio caminho.
    """
    if path.suffix.lower() == ".xlsb":
        return xlsb_to_xlsx(path, sheet_name)
    return path


# ===========================
#  ETAPA 1: INGESTÃO (pipeline)
# ===========================
def row_should_be_copied(ws, r, start_col, end_col):
    """Ignora linhas vazias e onde apenas a primeira coluna (B) está preenchida."""
    vals = [cellv(ws, r, c).strip() for c in range(start_col, end_col + 1)]
    if all(v == "" for v in vals):
        return False
    if vals[0] != "" and all(v == "" for v in vals[1:]):
        return False
    return True

def combine_two_excels(
    old_xlsx: Path,
    new_xlsx: Path,
    sheet_name: str,
    old_model_label: str,
    new_model_label: str,
    out_combined_path: Path,
    limit_to_z: bool = False,
) -> Path:
    """Abre as duas planilhas (mesmo nome), concatena, coluna A=MODELO, ignora col A de origem e pula linhas vazias/‘só B’."""
    if not old_xlsx.exists():
        raise FileNotFoundError(f"Arquivo (antigo) não encontrado: {old_xlsx}")
    if not new_xlsx.exists():
        raise FileNotFoundError(f"Arquivo (novo) não encontrado: {new_xlsx}")
    
    
    old_xlsx = ensure_xlsx(Path(old_xlsx), sheet_name)
    new_xlsx = ensure_xlsx(Path(new_xlsx), sheet_name)

    # aceita .xlsx/.xlsm; macros não são executadas; combinamos em um novo xlsx
    wb_old = load_workbook(old_xlsx, data_only=True)
    wb_new = load_workbook(new_xlsx, data_only=True)

    if sheet_name not in wb_old.sheetnames:
        raise ValueError(f'Aba "{sheet_name}" não encontrada em: {old_xlsx.name}')
    if sheet_name not in wb_new.sheetnames:
        raise ValueError(f'Aba "{sheet_name}" não encontrada em: {new_xlsx.name}')

    ws_old = wb_old[sheet_name]
    ws_new = wb_new[sheet_name]

    max_col_old = ws_old.max_column
    max_col_new = ws_new.max_column
    start_src_col = 2  # ignorar A
    max_col = max(max_col_old, max_col_new)
    #if limit_to_z:
    #    max_col = min(max_col, 26)  
    if max_col < start_src_col:
        raise ValueError("As planilhas parecem não ter colunas suficientes.")

    wb_out = Workbook()
    ws_out = wb_out.active

    # Cabeçalho: A = MODELO; B.. = cabeçalho original (a partir de B)
    ws_out.cell(row=1, column=1, value="MODELO")
    for c in range(start_src_col, max_col + 1):
        if c >= 16:  # 16 == coluna P (opcionais)
            old_header = str(ws_old.cell(row=7, column=c).value or "").strip()
            new_header = str(ws_new.cell(row=7, column=c).value or "").strip()
            if old_header and new_header and old_header != new_header:
                combined_header = f"{old_header}/{new_header}"
                #cell = ws_out.cell(row=1, column=c, value=combined_header)
                #Cor da fonte amarela
                #cell.font = Font(color="FFC000", bold=True)
            else:
                combined_header = old_header or new_header
        else:
            # para colunas antes de P, mantém o cabeçalho padrão (linha 1 do arquivo ANTIGO)
            combined_header = str(ws_old.cell(row=1, column=c).value or "").strip()
        ws_out.cell(row=1, column=c).value = combined_header

    out_r = 2
    # Copiar ANTIGO
    for r in range(2, ws_old.max_row + 1):
        if not row_should_be_copied(ws_old, r, start_src_col, max_col):
            continue
        ws_out.cell(row=out_r, column=1, value=old_model_label)
        for c in range(start_src_col, max_col + 1):
            ws_out.cell(row=out_r, column=c).value = ws_old.cell(row=r, column=c).value
        out_r += 1
    # Copiar NOVO
    for r in range(2, ws_new.max_row + 1):
        if not row_should_be_copied(ws_new, r, start_src_col, max_col):
            continue
        ws_out.cell(row=out_r, column=1, value=new_model_label)
        for c in range(start_src_col, max_col + 1):
            ws_out.cell(row=out_r, column=c).value = ws_new.cell(row=r, column=c).value
        out_r += 1

    wb_out.save(out_combined_path)
    return out_combined_path

# ===========================
#  ETAPA 2: PROCESSAMENTO
# ===========================
def process(in_path: Path, old_model: str, new_model: str, sheet_name=None, max_rows=None, limit_to_z=False):
    wb_in = load_workbook(in_path)
    ws_in = wb_in[sheet_name] if sheet_name else wb_in.active

    max_row = ws_in.max_row if max_rows is None else min(ws_in.max_row, int(max_rows))
    max_col = ws_in.max_column
    if limit_to_z:
        max_col = min(max_col, 26)

    wb = Workbook()
    ws = wb.active

    # 1) copiar cabeçalho original + dados
    for c in range(1, max_col+1):
        ws.cell(row=1, column=c).value = ws_in.cell(row=1, column=c).value
    out_r = 2

    # 2) remover duplicatas (C,E,F,J,K,L,M,N) com L normalizado
    key_cols = [COL_C, COL_E, COL_F, COL_J, COL_K, COL_L, COL_M, COL_N]

    def norm_for_key(wsref, r, c):
        if c == COL_L:
            v_num = celln(wsref, r, COL_L)
            if v_num is not None:
                return f"{round(v_num,6):.6f}"
        return clean_str(cellv(wsref, r, c))

    def row_key(r):
        return tuple(norm_for_key(ws_in, r, c) for c in key_cols)

    seen = defaultdict(list)
    for r in range(2, max_row+1):
        seen[row_key(r)].append(r)
    dup_rows = {r for rlist in seen.values() if len(rlist) > 1 for r in rlist}

    for r in range(2, max_row+1):
        if r in dup_rows:
            continue
        for c in range(1, max_col+1):
            ws.cell(row=out_r, column=c).value = ws_in.cell(row=r, column=c).value
        out_r += 1

    last_data_row = out_r - 1

    # 3) inserir coluna após L -> NOMINAL COUNT/LENGHT (3 casas)
    insert_at = COL_L + 1
    ws.insert_cols(insert_at, amount=1)
    ws.cell(row=1, column=insert_at, value="NOMINAL COUNT/LENGHT").font = BOLD

    _COL_M = COL_M + 1
    _COL_N = COL_N + 1
    _COL_O = COL_O + 1
    _COL_Q = COL_Q + 1

    for r in range(2, last_data_row+1):
        c_val = cellv(ws, r, COL_C).upper()
        d_val = cellv(ws, r, COL_D).upper()
        l_num = celln(ws, r, COL_L)
        value = None
        if l_num is not None:
            if c_val == "CABLE":
                value = l_num / 1.05
            elif c_val == "TAPE":
                if d_val == "CONTINUSLY LENGHT":
                    value = l_num / 3.0
                elif d_val == "SPIRALATO LENGHT":
                    value = l_num / 2.0
            elif c_val == "TUBE":
                value = l_num / 1.0
        if value is not None:
            c = ws.cell(row=r, column=insert_at, value=round(value, 3))
            c.number_format = "0.000"
        else:
            ws.cell(row=r, column=insert_at, value="-")

    # 4) inserir coluna após O: DELTA DRAWING (inicializar com "-")
    DELTA_COL = _COL_O + 1
    ws.insert_cols(DELTA_COL, amount=1)
    ws.cell(row=1, column=DELTA_COL, value="DELTA DRAWING").font = BOLD
    _COL_Q += 1
    for r in range(2, last_data_row + 1):
        ws.cell(row=r, column=DELTA_COL, value="-")
    ws.insert_cols(_COL_Q, amount=1)   # insere a coluna de STATUS na posição atual de _COL_Q (R)

    # 5) agrupamento por M e comparação
    groups = defaultdict(list)
    for r in range(2, last_data_row+1):
        groups[cellv(ws, r, _COL_M)].append(r)

    rows_to_delete = set()

    def compare_and_mark(r_old, r_new):
        diffs = []
        for c in [COL_C, COL_E, COL_F, COL_J, COL_K, COL_L, _COL_N]:
            if c == COL_L:
                if not equal_L(ws, r_old, r_new, COL_L):
                    diffs.append(COL_L)
            else:
                if clean_str(cellv(ws, r_old, c)) != clean_str(cellv(ws, r_new, c)):
                    diffs.append(c)

        # --- Regra de exclusão: única diferença é L e |Δ_M| <= 10 mm ---
        non_length_diffs = [c for c in diffs if c != COL_L]
        if not non_length_diffs:
            n_old = celln(ws, r_old, insert_at)
            n_new = celln(ws, r_new, insert_at)
            if n_old is None:
                n_old = parse_number_relaxed(ws.cell(row=r_old, column=insert_at).value)
            if n_new is None:
                n_new = parse_number_relaxed(ws.cell(row=r_new, column=insert_at).value)
            if (n_old is not None) and (n_new is not None):
                delta_mm_excl = (n_new - n_old) * 1000.0
                if abs(delta_mm_excl) <= 10.0005:
                    rows_to_delete.update((r_old, r_new))
                    return  # não marca nada; serão removidos depois

        # Casos especiais
        if set(diffs) == {COL_C}:
            msg = "Verificar classificação de componente"
            ws.cell(row=r_old, column=_COL_O, value=msg)
            ws.cell(row=r_new, column=_COL_O, value=msg)
            ws.cell(row=r_old, column=COL_C).fill = PINK
            ws.cell(row=r_new, column=COL_C).fill = PINK
            return

        if set(diffs) == {_COL_N}:
            msg = "Alteração de nome de opcional"
            ws.cell(row=r_old, column=_COL_O, value=msg)
            ws.cell(row=r_new, column=_COL_O, value=msg)
            ws.cell(row=r_old, column=_COL_N).fill = BLUE
            ws.cell(row=r_new, column=_COL_N).fill = BLUE
            return

        # Mensagens gerais
        msgs = []
        if COL_C in diffs:
            msgs.append("Alteração na classificação do componente")
        if _COL_N in diffs:
            msgs.append("Alteração de nome de opcional")
        if COL_J in diffs:
            msgs.append("Alteração do fornecedor")
        if COL_F in diffs:
            msgs.append("Alteração na especificação do componente")
        if COL_K in diffs:
            msgs.append("Alteração na unidade de medida")

        # Ajuste de comprimento (usa NOMINAL COUNT/LENGHT para o delta visível)
        if COL_L in diffs:
            n_old = celln(ws, r_old, insert_at)
            n_new = celln(ws, r_new, insert_at)
            if n_old is None:
                n_old = parse_number_relaxed(ws.cell(row=r_old, column=insert_at).value)
            if n_new is None:
                n_new = parse_number_relaxed(ws.cell(row=r_new, column=insert_at).value)

            if (n_old is not None) and (n_new is not None):
                delta_mm = (n_new - n_old) * 1000.0
                # só registra se |Δ| > 10 mm
                if abs(delta_mm) > 10.0:
                    msgs.append("Ajuste de comprimento")
                    cell_delta = ws.cell(row=r_new, column=DELTA_COL, value=f"{delta_mm:.0f} mm")
                    cell_delta.font = Font(bold=True, color=("008000" if delta_mm >= 0 else "FF0000"))

        if (COL_E in diffs) and (COL_J not in diffs):
            msgs.append("Alteração de PartNumber")

        if msgs:
            joined = " / ".join(msgs)
            ws.cell(row=r_old, column=_COL_O, value=joined)
            ws.cell(row=r_new, column=_COL_O, value=joined)

        for c in diffs:
            ws.cell(row=r_old, column=c).fill = YELL
            ws.cell(row=r_new, column=c).fill = YELL

    # Pareamento por grupo (ordem de aparição)
    for mval, rlist in groups.items():
        q_old, q_new = deque(), deque()
        for r in rlist:
            a = cellv(ws, r, COL_A)
            if a == old_model: q_old.append(r)
            elif a == new_model: q_new.append(r)
        while q_old or q_new:
            r2 = q_old.popleft() if q_old else None
            r3 = q_new.popleft() if q_new else None
            if r2 and r3:
                compare_and_mark(r2, r3)
            elif r2 and not r3:
                ws.cell(row=r2, column=_COL_O, value="Eliminado")
            elif r3 and not r2:
                ws.cell(row=r3, column=_COL_O, value="Novo")

    # Remover as duplas marcadas
    if rows_to_delete:
        for r in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(r, 1)
        last_data_row = ws.max_row

    # 6) status em Q
    for r in range(2, last_data_row+1):
        a = cellv(ws, r, COL_A)
        if a == old_model:
            cell = ws.cell(row=r, column=_COL_Q); cell.value = "REMOVIDO"
            cell.fill = RED; cell.font = WHITE_BOLD
        elif a == new_model:
            cell = ws.cell(row=r, column=_COL_Q); cell.value = "ADICIONADO"
            cell.fill = GREEN; cell.font = BLACK_BOLD

    # 7) pintar N conforme O
    for r in range(2, last_data_row+1):
        otext = cellv(ws, r, _COL_O)
        if otext == "Novo":
            ws.cell(row=r, column=_COL_N).fill = GREEN
        elif otext == "Eliminado":
            ws.cell(row=r, column=_COL_N).fill = RED

    # 8) varrer colunas R+ até vazia (substituir "X" por ±L) — com limite congelado
    start_col = _COL_Q + 1
    col = start_col
    max_cols_scan = ws.max_column
    while col <= max_cols_scan:
        has_content = any(cellv(ws, r, col).strip() != "" for r in range(2, last_data_row + 1))
        if not has_content:
            break
        for r in range(2, last_data_row + 1):
            val = cellv(ws, r, col).strip().upper()
            if val == "X":
                a_val = cellv(ws, r, COL_A)
                l_num = celln(ws, r, COL_L)
                if l_num is None:
                    ws.cell(row=r, column=col, value=None)
                else:
                    if a_val == old_model:
                        new_val = -abs(l_num)
                    elif a_val == new_model:
                        new_val = abs(l_num)
                    else:
                        continue
                    c = ws.cell(row=r, column=col, value=round(new_val, 3))
                    c.number_format = "0.000"
        col += 1

    # 9) cabeçalho extra e títulos na linha 7
    ws.insert_rows(1, amount=6)
    ws.cell(row=1, column=1, value="DELTA BOM").font = Font(name="Arial", size=16, bold=True, color="000080")
    ws.cell(row=2, column=1, value="RESPONSÁVEL:").font = BOLD
    ws.cell(row=3, column=1, value="PROJETO:").font = BOLD
    ws.cell(row=4, column=1, value="FAMÍLIA:").font = BOLD
    ws.cell(row=5, column=1, value="CCD/ODM:").font = BOLD

    ws.cell(row=2, column=3, value="DATA:").font = BOLD
    ws.cell(row=4, column=3, value="NÍVEL INICIAL").font = BOLD
    ws.cell(row=5, column=3, value="NÍVEL FINAL").font = BOLD

    today = datetime.today().date()
    c = ws.cell(row=2, column=4, value=today); c.font = BOLD; c.number_format = "DD/MM/YYYY"
    ws.cell(row=4, column=4, value=old_model).font = BOLD
    ws.cell(row=5, column=4, value=new_model).font = BOLD

    # Títulos da tabela na linha 7
    ws.cell(row=7, column=2,  value="ITEM")
    ws.cell(row=7, column=3,  value="COMPONENT FAMILY")
    ws.cell(row=7, column=4,  value="CONFIGURATION")
    ws.cell(row=7, column=5,  value="CÓDIGO DE REFERÊNCIA")
    ws.cell(row=7, column=6,  value="DESCRIÇÃO")
    ws.cell(row=7, column=7,  value="CÓDIGO DO FORNECEDOR")
    ws.cell(row=7, column=8,  value="CÓDIGO DO CLIENTE")
    ws.cell(row=7, column=9,  value="CÓDIGO FCA-JBT")
    ws.cell(row=7, column=10, value="FORNECEDOR")
    ws.cell(row=7, column=11, value="U.M.")
    ws.cell(row=7, column=12, value="TOTAL COUNT/LENGHT")
    # 13 (M) é NOMINAL COUNT/LENGHT (inserida)
    ws.cell(row=7, column=14, value="TAKE OUT")
    ws.cell(row=7, column=15, value="OPCIONAL REFERÊNCIA")
    ws.cell(row=7, column=16, value="NOTAS")
    ws.cell(row=7, column=18, value="STATUS")
    
    # Legendas
    ws.cell(row=2, column=14, value="LEGENDA:").font = BOLD
    c = ws.cell(row=3, column=14, value="NOVO");       c.fill = GREEN; c.font = BOLD
    c = ws.cell(row=4, column=14, value="REMOVIDO");   c.fill = RED;   c.font = BOLD
    c = ws.cell(row=5, column=14, value="MODIFICADO"); c.fill = YELL;  c.font = BOLD
    c = ws.cell(row=4, column=15, value="ALTERAÇÃO NA COMPONENT FAMILY"); c.fill = PINK; c.font = BOLD
    c = ws.cell(row=5, column=15, value="MUDANÇA DE OPCIONAL");            c.fill = BLUE; c.font = BOLD

    # Pintar linha 7
    max_col_after = ws.max_column
    for c in range(1, max_col_after + 1):
        cell = ws.cell(row=7, column=c)
        cell.fill = DARK_BLUE
        cell.font = WHITE_BOLD
        
    first_optional_col = _COL_Q + 1  # ou STATUS_COL + 1
    for c in range(first_optional_col, ws.max_column + 1):
        txt = ws.cell(row=7, column=c).value
        if isinstance(txt, str) and '/' in txt:
            # Se a partir da coluna Q, ele identificar que existe "/" ele pinta o texto de amarelo, se não, ele mantém branco
            ws.cell(row=7, column=c).font = Font(color="FFFF00", bold=True)
        
    auto_width(ws)

    out_path = in_path.with_name(in_path.stem + "_PROCESSADO.xlsx")
    wb.save(out_path)
    return out_path

# ===========================
#  Worker (QThread)
# ===========================
class PipelineWorker(QThread):
    progress = Signal(str)
    done = Signal(str)
    error = Signal(str)

    def __init__(self, old_path, new_path, sheet, old_model, new_model, limit_to_z, outdir: Path):
        super().__init__()
        self.old_path = Path(old_path)
        self.new_path = Path(new_path)
        self.sheet = sheet
        self.old_model = old_model
        self.new_model = new_model
        self.limit_to_z = limit_to_z
        self.outdir = Path(outdir)

    def run(self):
        try:
            self.progress.emit("[1/2] Combinando arquivos (ignorando col A e pulando linhas vazias/‘só B’)…")
            self.outdir.mkdir(parents=True, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            combined_path = self.outdir / f"DELTA_COMBINADO_{ts}.xlsx"

            combined = combine_two_excels(
                old_xlsx=self.old_path,
                new_xlsx=self.new_path,
                sheet_name=self.sheet,
                old_model_label=self.old_model,
                new_model_label=self.new_model,
                out_combined_path=combined_path,
                limit_to_z=self.limit_to_z,
            )
            self.progress.emit(f"✅ Combinado: {combined}")

            self.progress.emit("[2/2] Processando regras…")
            final_path = process(
                in_path=combined,
                old_model=self.old_model,
                new_model=self.new_model,
                sheet_name=None,
                max_rows=None,
                limit_to_z=self.limit_to_z
            )
            self.done.emit(str(final_path))
        except Exception as e:
            self.error.emit(str(e))

# ===========================
#  GUI
# ===========================
class DeltaBOMGUI(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Delta BOM")
        self.resize(900, 700)
        
        layout = QVBoxLayout()

        # Header com título + modo noturno
        header = QHBoxLayout()
        title = QLabel("Delta BOM – Pipeline")
        title.setStyleSheet("font-weight: 700; font-size: 16pt;")
        header.addWidget(title)
        header.addStretch(1)
        # --- checkbox já existente (ou crie aqui se não existir) ---
        # se você JÁ criou self.chk_night antes, não recrie; só garanta que ele exista aqui
        # e que a conexão do sinal venha APÓS a criação do logo_label.
        self.chk_night = QCheckBox("Dark Mode")
        self.chk_night.setChecked(False)
        header.addWidget(self.chk_night)

        layout.addLayout(header)


        # Arquivo ANTIGO
        layout.addWidget(QLabel("OLD File (.xlsx/.xlsm):"))
        row1 = QHBoxLayout()
        self.ent_old = QLineEdit()
        btn_old = QPushButton("Search…")
        btn_old.clicked.connect(self.pick_old)
        row1.addWidget(self.ent_old)
        row1.addWidget(btn_old)
        layout.addLayout(row1)

        # Arquivo NOVO
        layout.addWidget(QLabel("NEW File (.xlsx/.xlsm):"))
        row2 = QHBoxLayout()
        self.ent_new = QLineEdit()
        btn_new = QPushButton("Search…")
        btn_new.clicked.connect(self.pick_new)
        row2.addWidget(self.ent_new)
        row2.addWidget(btn_new)
        layout.addLayout(row2)

        # Aba e modelos
        row3 = QHBoxLayout()
        row3.addWidget(QLabel("Tab Name:"))
        self.ent_sheet = QLineEdit()
        row3.addWidget(self.ent_sheet)
        layout.addLayout(row3)

        row4 = QHBoxLayout()
        row4.addWidget(QLabel("OLD Model (ex.: X2):"))
        self.ent_old_model = QLineEdit()
        row4.addWidget(self.ent_old_model)

        row4.addWidget(QLabel("NEW Model (ex.: X3):"))
        self.ent_new_model = QLineEdit()
        row4.addWidget(self.ent_new_model)
        layout.addLayout(row4)

        # Pasta de saída
        layout.addWidget(QLabel("Output folder:"))
        row5 = QHBoxLayout()
        self.ent_outdir = QLineEdit(".")
        btn_out = QPushButton("Search…")
        btn_out.clicked.connect(self.pick_outdir)
        row5.addWidget(self.ent_outdir)
        row5.addWidget(btn_out)
        layout.addLayout(row5)

        # Opções
        # row6 = QHBoxLayout()
        # self.chk_limit_z = QCheckBox("Limitar às colunas A..Z")
        # self.chk_limit_z.setChecked(True)
        # row6.addWidget(self.chk_limit_z)
        # row6.addStretch(1)
        # layout.addLayout(row6)

        # Botão executar
        self.btn_run = QPushButton("Execute pipeline")
        self.btn_run.clicked.connect(self.run_pipeline)
        layout.addWidget(self.btn_run)

        # Log
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setStyleSheet("""
            QTextEdit {
                background: transparent;
                border: none;
            }
        """)
        self.log.setAttribute(Qt.WA_TranslucentBackground, True)
        self.log.setFrameStyle(0)
        layout.addWidget(self.log)

        # Rodapé créditos
        footer = QLabel(
            "Responsible: Christian Drumond Marques - Systems Engineer "
        )
        layout.addWidget(footer, alignment=Qt.AlignLeft)

        
        # IMAGEM DE FUNDO
        self.bg_label = QLabel(self)
        self.bg_label.setScaledContents(True)
        self.bg_label.resize(307, 204)
        self.bg_label.lower()
        # caminhos das imagens (ajuste os nomes)
        self.bg_img_dark  = "ChicoteCarro2.png"
        self.bg_img_light = "ChicoteCarro.png"
        # aplica a imagem inicial de acordo com o checkbox (modo noturno)
        initial_pix = QPixmap(self.bg_img_dark if self.chk_night.isChecked() else self.bg_img_light)
        self.bg_label.setPixmap(initial_pix)
        # posição inicial (o seu resizeEvent já realinha no canto)
        self.bg_label.move(self.width() - 210, self.height() - 210)
        # Mantém a imagem atrás dos demais widgets
        self.bg_label.lower()

        # Conecte o sinal SÓ agora, após logo_label existir
        self.chk_night.toggled.connect(self.apply_theme)
        # Aplique o tema uma vez para setar o logo inicial
        self.apply_theme(self.chk_night.isChecked())

        self.setLayout(layout)

    # ---- tema ----
    def apply_theme(self, checked: bool):
        app = QApplication.instance()
        app.setStyleSheet(DARK_QSS if checked else LIGHT_QSS)
        
        # troca a imagem conforme o tema
        from PySide6.QtGui import QPixmap
        pix = QPixmap(self.bg_img_dark if checked else self.bg_img_light)
        self.bg_label.setPixmap(pix)
        self.bg_label.setScaledContents(True)  # mantém comportamento atual
        

    # ---- diálogos ----
    def pick_old(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select OLD Excel", "", "Excel Files (*.xlsx *.xlsm *.xlsb);;All Files (*)"
        )
        if path:
            self.ent_old.setText(path)

    def pick_new(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select NEW Excel", "", "Excel Files (*.xlsx *.xlsm *.xlsb);;All Files (*)"
        )
        if path:
            self.ent_new.setText(path)

    def pick_outdir(self):
        path = QFileDialog.getExistingDirectory(self, "Select the output folder.", "")
        if path:
            self.ent_outdir.setText(path)

    # ---- log ----
    def append_log(self, msg: str):
        self.log.append(msg)

    # ---- pipeline ----
    def run_pipeline(self):
        old = self.ent_old.text().strip().strip('"')
        new = self.ent_new.text().strip().strip('"')
        sheet = self.ent_sheet.text().strip()
        old_model = self.ent_old_model.text().strip()
        new_model = self.ent_new_model.text().strip()
        outdir = self.ent_outdir.text().strip() or "."

        if not (old and new and sheet and old_model and new_model):
            self.append_log("⚠️ Fill in all the fields..")
            return

        if not Path(old).exists() or not Path(new).exists():
            self.append_log("⚠️ File path invalid.")
            return

        self.btn_run.setEnabled(False)
        self.append_log("⏳ Iniciando…")

        self.worker = PipelineWorker(
            old_path=old,
            new_path=new,
            sheet=sheet,
            old_model=old_model,
            new_model=new_model,
            limit_to_z=False,
            outdir=Path(outdir)
        )
        self.worker.progress.connect(self.append_log)
        self.worker.done.connect(lambda p: (self.append_log(f"✅ Completed! Final file: {p}"), self.btn_run.setEnabled(True)))
        self.worker.error.connect(lambda msg: (self.append_log(f"❌ ERRO: {msg}"), self.btn_run.setEnabled(True)))
        self.worker.start()
        
    # Reposiciona a imagem quando a janela é redimensionada
    def resizeEvent(self, event):
        super().resizeEvent(event)
        if hasattr(self, "bg_label"):
            w, h = self.bg_label.width(), self.bg_label.height()
            # margem de 10 px do canto inferior direito
            self.bg_label.move(self.width() - w - 10, self.height() - h - 10)
        
# ===========================
#  MAIN
# ===========================
if __name__ == "__main__":
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    app = QApplication(sys.argv)
    app.setStyleSheet(LIGHT_QSS)  # começa no modo noturno; o checkbox alterna
    window = DeltaBOMGUI()
    window.setWindowIcon(QIcon("icone_multi_cropped.ico"))
    window.show()
    sys.exit(app.exec())
