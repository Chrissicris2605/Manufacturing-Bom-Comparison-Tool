"""Excel report generation for the public BOM comparison demo."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models import STATUS_STYLES


def _autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def _style_header(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def generate_excel_report(delta_df: pd.DataFrame, output_path: str | Path) -> Path:
    """Generate a formatted Excel report with summary and detailed delta sheets."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    summary_df = (
        delta_df["status"]
        .value_counts()
        .reindex(["Added", "Removed", "Modified", "Unchanged"], fill_value=0)
        .rename_axis("status")
        .reset_index(name="count")
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        delta_df.to_excel(writer, index=False, sheet_name="Delta Report")

    workbook = load_workbook(output_path)

    summary_sheet = workbook["Summary"]
    delta_sheet = workbook["Delta Report"]

    for worksheet in (summary_sheet, delta_sheet):
        _style_header(worksheet)
        _autosize_columns(worksheet)
        worksheet.freeze_panes = "A2"

    # Color status cells in both sheets.
    for worksheet in (summary_sheet, delta_sheet):
        status_column = 1
        for row in range(2, worksheet.max_row + 1):
            status = worksheet.cell(row=row, column=status_column).value
            if status in STATUS_STYLES:
                fill = PatternFill("solid", fgColor=STATUS_STYLES[status].fill_color)
                worksheet.cell(row=row, column=status_column).fill = fill
                worksheet.cell(row=row, column=status_column).font = Font(bold=True)

    workbook.save(output_path)
    return output_path
